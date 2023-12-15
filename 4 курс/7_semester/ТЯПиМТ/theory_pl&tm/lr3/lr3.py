import dearpygui.dearpygui as dpg
import regex as re
from __init__ import initialize, select_path
from openpyxl import load_workbook
import openpyxl
from enum import IntEnum
from anytree import Node, RenderTree


class LL:
    table: openpyxl.worksheet.worksheet.Worksheet
    buffer: str
    root: Node
    current_node: Node

    def __init__(self):
        self.buffer = ""
        self.root = Node('root')
        self.current_node = self.root

    class Column(IntEnum):
        LEFT = 1
        TERMS = 2
        JUMP = 3
        ACCEPT = 4
        STACK = 5
        RETURN = 6
        ERROR = 7
        ACTION = 8

    def open_parse_table(self, file_parse_table: str):
        try:
            self.table = load_workbook(filename=file_parse_table, read_only=True)['parse_table']
        except BaseException as err:
            raise FileNotFoundError(err)

    # region Generate parse table
    class ParseTable:
        wb: openpyxl.Workbook
        ws: openpyxl.worksheet.worksheet.Worksheet
        parse_table: openpyxl.worksheet.worksheet.Worksheet
        actions: openpyxl.worksheet.worksheet.Worksheet
        dict_LL: dict
        dict_M: dict
        start_col: int
        follow_col: int
        num_rows: int

        def __init__(self):
            self.dict_LL = dict()
            self.dict_M = dict()
            self.wb = openpyxl.Workbook()
            self.parse_table = self.wb.create_sheet("parse_table", 0)
            self.parse_table.append(
                ('НЕТЕРМИНАЛЫ', "terminals", "jump", "accept", "stack", "return", "error", "action"))
            self.ws = self.wb.create_sheet("temp_list", 1)
            self.ws.append(('ПРАВИЛО', "START0"))
            self.actions = self.wb.create_sheet("actions", 2)
            self.actions.append(('ПРАВИЛО', "action"))

        def generate_parse_table(self, file_grammar: str):
            self.parse_raw_rules(file_grammar)
            self.find_start_nodes()
            self.find_follow_nodes()
            self.find_direction_nodes()
            self.find_jumps()
            self.put_actions()
            self.wb.save(f"{file_grammar}.xlsx")

        @staticmethod
        def is_nonterm(s):
            return s[0].isupper()

        def parse_raw_rules(self, file_grammar):
            # print("parse raw rules")
            try:
                with open(file_grammar) as file:
                    index = 0
                    global_index = 1
                    for row in file.readlines():
                        if len(list_splitted := row.split(" ->")) == 2:
                            left, right = list_splitted
                            action = ''
                            if len(left_list := left.split(" ")) == 2:
                                left, action = left_list
                            for rule in right.split('|'):
                                global_index += 1
                                self.actions.append((f"{left} -> ", action))
                            # left_term, action = left.split(" ")
                            rules = {}
                            check_list = []
                            for rule in right.split('|'):
                                rule = rule.strip()
                                nodes = rule.split(" ")
                                rule = re.sub(r'\s<.*?>', '', rule)
                                for node in nodes:
                                    global_index += 1
                                    value = f"{node}"
                                    if node[0] == '<':
                                        global_index -= 1
                                        action = node
                                        self.actions.cell(row=global_index, column=2, value=action)
                                        # print(value, action)
                                    else:
                                        self.actions.append((value, ""))
                                rules.update({index: rule})
                                state0 = "∅" if LL.ParseTable.is_nonterm(rule) \
                                    else rule.split(" ")[0]
                                check_list.append(state0)
                                # print("left tert: ", left_term)
                                self.ws.append((f"{left} -> {rule}", state0))
                                index += 1
                            LL.ParseTable.check_duplicates(list(filter('∅'.__ne__, check_list)),
                                                           left)
                            self.dict_LL[left] = rules
            except BaseException as err:
                raise FileExistsError(err)

        def put_actions(self):
            for row in range(2, self.actions.max_row + 1):
                value = self.actions.cell(row=row, column=2).value
                self.parse_table.cell(row=row, column=LL.Column.ACTION, value=value)

        @staticmethod
        def is_equal_cols(rows: int, worksheet: openpyxl.worksheet.worksheet.Worksheet,
                          compare_col1: int, compare_col2: int) -> bool:
            for row in range(2, rows + 1):
                if worksheet.cell(row=row, column=compare_col1).value \
                        != worksheet.cell(row=row, column=compare_col2).value:
                    return False
            return True

        @staticmethod
        def check_duplicates(check_list: list, key: str, stage: str = 'starting'):
            for item in check_list:
                if check_list.count(item) > 1:
                    check_list = " ".join(
                        [f"{i}\n" if (index + 1) % 10 == 0 else i for index, i in enumerate(check_list)])
                    raise GeneratorExit(f"Duplicates were encountered during the generation of {stage} symbols:"
                                        f"\n[{check_list}]"
                                        f"\nduplicated symbol: {item}"
                                        f"\noccurred in left term: {key}")

        def find_start_nodes(self):
            # print("find_start_nodes")
            self.start_col = 2
            self.num_rows = 0
            while True:
                self.ws.cell(row=1, column=self.start_col + 1, value=f"START{self.start_col - 1}")
                for key in self.dict_LL.keys():
                    check_list = []
                    for index, rule in self.dict_LL[key].items():
                        new_starts = []
                        if LL.ParseTable.is_nonterm(rule):
                            key_start = rule.split(" ")[0]
                            for start_index_row in self.dict_LL[key_start].keys():
                                new_starts.append(self.ws.cell(row=start_index_row + 2, column=self.start_col).value)
                            new_starts = list(filter('∅'.__ne__, new_starts))
                            if len(new_starts) == 0:
                                new_starts.append('∅')
                        else:
                            new_starts = [self.ws.cell(row=index + 2, column=self.start_col).value]
                        check_list += new_starts
                        self.ws.cell(row=index + 2, column=self.start_col + 1, value=" ".join(new_starts))
                        self.num_rows = index + 2
                    check_list = list(filter('∅'.__ne__, check_list))
                    LL.ParseTable.check_duplicates(check_list, key)
                if LL.ParseTable.is_equal_cols(self.num_rows, self.ws, self.start_col, self.start_col + 1):
                    break
                self.start_col += 1

        def get_row_follow_node(self, key, operation=min):
            return operation(self.dict_LL[key].keys()) + 2

        def get_follow_nodes(self, key, col) -> list:
            raw_follows = self.ws.cell(row=self.get_row_follow_node(key, min), column=col).value
            return raw_follows.split(" ") if raw_follows else []

        def set_follow_nodes(self, key, col, follows: list):
            self.ws.cell(row=self.get_row_follow_node(key, min), column=col, value=" ".join(follows))

        def review_next(self, key, nodes, index_node, follows, base_col):
            next_node = ''
            try:
                next_node = nodes[index_node + 1]
                # итерация по правилам следующего узла для поиска его символов-предсшественников
                for start_index_row in self.dict_LL[next_node].keys():
                    start_terms: list = self.ws.cell(row=start_index_row + 2, column=self.start_col).value.split(" ")
                    if 'e' in start_terms:
                        start_terms.remove('e')
                        follows += self.review_next(key, nodes, index_node + 1, follows, base_col)
                    else:
                        follows += start_terms
            except IndexError as _:
                follows += self.get_follow_nodes(key, base_col)
            except KeyError as _:
                follows.append(next_node)
            follows = list(set(follows))
            follows.sort()
            return follows

        def put_follows(self, col: int):
            self.ws.cell(row=1, column=col, value=f"FOLLOWS{col - self.start_col - 1}")
            # итерация по нетерминалам
            for key in self.dict_LL.keys():
                # print("key: ", key)
                # итерация по правилам нетерминала
                for index_rule, rule in self.dict_LL[key].items():
                    nodes = rule.split(" ")
                    # print("nodes: ", nodes)
                    # итерация по терминальным и нетерминальным узлам
                    for index_node, node in enumerate(nodes):
                        # для каждого нетерминала определяем последующие узлы
                        if LL.ParseTable.is_nonterm(node):
                            follows = self.get_follow_nodes(node, col)
                            follows = self.review_next(key, nodes, index_node, follows, col)
                            self.set_follow_nodes(node, col, follows)

        def find_follow_nodes(self):
            # print("find_follow_nodes")
            self.start_col += 1
            self.follow_col = self.start_col + 1
            for key in self.dict_LL.keys():
                self.ws.merge_cells(start_row=self.get_row_follow_node(key, min), start_column=self.follow_col,
                                    end_row=self.get_row_follow_node(key, max), end_column=self.follow_col)
            self.ws.cell(row=2, column=self.follow_col, value="⊥")
            self.put_follows(self.follow_col)
            self.follow_col += 1
            count_interations = 0
            while True:
                old_value = []
                for key in self.dict_LL.keys():
                    self.ws.merge_cells(start_row=self.get_row_follow_node(key, min), start_column=self.follow_col,
                                        end_row=self.get_row_follow_node(key, max), end_column=self.follow_col)
                    old_value = self.get_follow_nodes(key, self.follow_col - 1)
                    self.set_follow_nodes(key, self.follow_col, old_value)
                self.put_follows(self.follow_col)
                if LL.ParseTable.is_equal_cols(self.num_rows, self.ws, self.follow_col - 1, self.follow_col):
                    break
                self.follow_col += 1
                count_interations += 1
                if count_interations > 20:
                    raise StopIteration("Oops, the algorithm entered an infinite loop"
                                        f"follows: {old_value}")

        def next_starts(self, key, starts, nodes, index_node):
            try:
                next_node = nodes[index_node + 1]
                for start_index_row in self.dict_LL[next_node].keys():
                    starts += self.ws.cell(row=start_index_row + 2, column=self.start_col).value.split(" ")
                    if 'e' in starts:
                        starts.remove("e")
                        starts = self.next_starts(key, starts, nodes, index_node + 1)
            except IndexError as _:
                starts += self.get_follow_nodes(key, self.follow_col)
            except KeyError as _:
                starts.append(next_node)
            return starts

        @staticmethod
        def is_last_node(current_index, rule_nodes):
            try:
                if rule_nodes[current_index + 1]:
                    return False
            except IndexError as _:
                return True

        def find_direction_nodes(self):
            # print("find_direction_nodes")
            self.ws.cell(row=1, column=self.follow_col + 1, value="DIRECTIONS")
            index_term = 1
            for key in self.dict_LL.keys():
                index_term = self.find_left_direction_nodes(key, index_term)
                index_term = self.find_right_direction_nodes(key, index_term, index_term - 1)

        def find_left_direction_nodes(self, key, index_term):
            check_list = []
            for index in self.dict_LL[key].keys():
                index_term += 1
                direction_terms: list = self.ws.cell(row=index + 2, column=self.start_col).value.split(" ")
                if 'e' in direction_terms:
                    direction_terms.remove("e")
                    follow_e = self.get_follow_nodes(key, self.follow_col)
                    direction_terms += follow_e
                check_list += direction_terms
                self.ws.cell(row=index + 2, column=self.follow_col + 1, value=" ".join(direction_terms))
                self.dict_M.update({(key, index_term): {}})
                self.parse_table.append(
                    ("left: " + key, " ".join(direction_terms), "", "False", "False", "False", "False"))
            check_list = list(filter('⊥'.__ne__, check_list))
            LL.ParseTable.check_duplicates(check_list, key, 'direction')
            self.parse_table.cell(row=index_term, column=LL.Column.ERROR, value="True")
            return index_term

        def find_right_direction_nodes(self, key, index_term, last_left_term):
            for index_rule, rule in self.dict_LL[key].items():
                nodes = rule.split(" ")
                # итерация по терминальным и нетерминальным узлам
                for index_node, node in enumerate(nodes):
                    index_term += 1
                    # для каждого нетерминала определяем направляющие узлы
                    accept = "False"
                    stack = "False"
                    # has_action = True
                    if LL.ParseTable.is_nonterm(node):
                        stack = str(not LL.ParseTable.is_last_node(index_node, nodes))
                        starts = []
                        for start_index_row in self.dict_LL[node].keys():
                            starts += self.ws.cell(row=start_index_row + 2, column=self.start_col).value.split(" ")
                            if 'e' in starts:
                                starts.remove("e")
                                starts = self.next_starts(key, starts, nodes, index_node)
                    elif node == 'e':
                        starts = self.get_follow_nodes(key, self.follow_col)
                    else:
                        starts = [node]
                        accept = "True"

                    global_index_rule = index_rule - min(self.dict_LL[key].keys())
                    self.dict_M[(key, last_left_term - len(self.dict_LL[key]) + 2 + global_index_rule)].update(
                        {index_term: node})
                    # print(self.dict_M)
                    self.parse_table.append(("right: " + node, " ".join(starts), "", accept, stack, "False", "True"))
            return index_term

        def find_jumps(self):
            # print("find_jumps")
            for key, values in self.dict_M.items():
                self.parse_table.cell(row=key[1], column=LL.Column.JUMP, value=min(values.keys()))
                for index_node, node in values.items():
                    if LL.ParseTable.is_nonterm(node):
                        for root_key in self.dict_M.keys():
                            if root_key[0] == node:
                                self.parse_table.cell(row=index_node, column=LL.Column.JUMP, value=root_key[1])
                                break
                    else:
                        try:
                            next_node = values[index_node + 1]
                            self.parse_table.cell(row=index_node, column=LL.Column.JUMP, value=index_node + 1)
                        except KeyError as _:
                            self.parse_table.cell(row=index_node, column=LL.Column.JUMP, value=0)
                            self.parse_table.cell(row=index_node, column=LL.Column.RETURN, value="True")

    # endregion

    def parse_value(self, row, name_col: Column):
        return self.table.cell(row=row, column=name_col).value

    @staticmethod
    def what_is_node(node: Node):
        is_node = 'node'
        for child in node.children:
            if child.name[0] == '{' and child.name[-1] == '}':
                is_node = 'implementation'
            elif child.name == ';':
                is_node = 'prototype'
            elif child.name == '}':
                is_node = 'namespace'
        return is_node

    @staticmethod
    def is_prototype(token):
        return token == ';'

    @staticmethod
    def is_implementation(token):
        return token[0] == '{' and token[-1] == '}'

    @staticmethod
    def check_ids(node: Node, check_end=is_prototype):
        """
            Проверка на:
            - уникальность идентификаторов, если они есть
        """
        ids = []
        for type_id in node.children:
            if check_end(type_id.name):
                break
            id_node = type_id.children
            if id_node:
                ids.append(type_id.children[0].name)
        for id_name in ids:
            if ids.count(id_name) > 1:
                raise NameError(f"Duplicated id: {id_name}")

    @staticmethod
    def check_implementation(node: Node, implementation: list):
        """
            Проверка на:
            - уникальность идентификаторов с списке параметров
            - отсутствие конфликтов в перегрузках функций
        """
        LL.check_ids(node, LL.is_implementation)
        # идентификатор функции и типы данных, перечисленные в параметрах функции
        func_args = [node.name]
        # перебираем типы данных параметров функции (реализации)
        for type_id in node.children:
            # узел '{}' был заглушкой для определения того, что родительский узел это реализация функции
            if LL.is_implementation(type_id.name):
                break
            func_args.append(type_id.name)
        # пробуем добавить список [id_функции1, тип2, тип1, ...]
        # используется полное соответствие, поэтому, если уже была перегрузка [id_функции1, тип1, тип2, ...],
        # то добавляемый список релевантный
        for item in implementation:
            if func_args == item:
                raise TypeError(f"Duplication in function overloads. Function: {node.name}")
        return [func_args]

    @staticmethod
    def check_semantics(node: Node):
        """
            Проверка семантики:
            - Несовпадение имён пространств и имён функций (прототип или реализация)
            - Уникальность идентификаторов в рамках одной функции (прототип или реализация)
            - Уникальность перегрузок реализаций (проверяется имя функции и типы списка параметров)
        """
        names = {'implementation': [],
                 'prototype': [],
                 'namespace': [],
                 'node': []}
        implementation = list()
        for child in node.children:
            names[LL.what_is_node(child)] += [child.name]
            match LL.what_is_node(child):
                case 'implementation':
                    implementation += LL.check_implementation(child, implementation)
                case 'namespace':
                    LL.check_semantics(child)
                case 'prototype':
                    LL.check_ids(child)
        for name in names['namespace']:
            if name in names['prototype'] + names['implementation']:
                raise NameError(f"Duplicated name: {name}")

    def apply_func(self, token: str, action: str):
        match action:
            case '<A1>':
                self.buffer += token
            case '<A2>':
                self.buffer = ''
            case '<APPEND>':
                self.current_node = Node(self.buffer, parent=self.current_node)
                self.buffer = ''
            case '<NAMESPASE>':
                for child in self.current_node.children:
                    if self.buffer == child.name:
                        child.children = child.children[:-1]
                        self.current_node = child
                        self.buffer = ''
                        return
                self.current_node = Node(self.buffer, parent=self.current_node)
                self.buffer = ''
            case '<ADD>':
                Node(self.buffer, parent=self.current_node)
                self.buffer = ''
            case '<\APPEND>':
                self.current_node = self.current_node.parent
            case '<RW>':
                Node(token, parent=self.current_node)
                self.buffer = ''
            case '<RW_DOWN>':
                Node(token, parent=self.current_node)
                self.current_node = self.current_node.parent
                self.buffer = ''
            case '<APPEND_DOWN>':
                self.buffer += token
                Node(self.buffer, parent=self.current_node)
                self.buffer = ''
                self.current_node = self.current_node.parent

    def analyze(self, input_string: str):
        input_string += '⊥'
        i = 2
        k = 0
        Stack = [0]
        while True:
            terms_str = "|".join(self.parse_value(i, LL.Column.TERMS).split(" "))
            terms = re.compile(terms_str)
            if match := terms.match(input_string, pos=k):
                match = match[0]
                len_shift = len(match)
                self.apply_func(match, self.parse_value(i, LL.Column.ACTION))
                if self.parse_value(i, LL.Column.ACCEPT) == 'True':
                    # print("ACCEPTED")
                    # print("match: ", match)
                    k += len_shift
                if self.parse_value(i, LL.Column.STACK) == "True":
                    # print("STACKED")
                    Stack.append(i + 1)
                if self.parse_value(i, LL.Column.RETURN) == "True":
                    # print("RETURNED")
                    i = Stack.pop()
                    if i == 0:
                        break
                    else:
                        continue
                else:  # LL.Column.JUMP
                    i = self.parse_value(i, LL.Column.JUMP)
                    # print("JUMP to ", i)
            elif self.parse_value(i, LL.Column.ERROR) == "False":
                i += 1
            else:
                break
        if len(Stack) == 0 and match == '⊥':
            print("tree:\n")
            for pre, fill, node in RenderTree(self.root):
                print("%s%s" % (pre, node.name))
            self.check_semantics(self.root)
            return f"SUCCESS PARSED!"
        else:
            raise SyntaxError(f"FAILED PARSED! at: {k}\n->{input_string[k:]}")


def main():
    dpg.show_item('Analyzing')
    input_data = get_input_data()
    data_with_numering = '\n'.join([f'{index}\t{row}'
                                    for index, row
                                    in enumerate(input_data.split('\n'), 1)])
    dpg.set_value('input data', value=data_with_numering)
    dpg.set_value('test', value="")
    try:
        engine: LL = LL()
        engine.open_parse_table(dpg.get_value('file_grammar'))
        message = engine.analyze(input_data)
        dpg.configure_item('test', default_value=f"{message}\n{input_data}", color=(0, 255, 0, 255))
    except BaseException as err:
        dpg.configure_item('test', default_value=f"Exception error during analyzing:\n{err}", color=(255, 0, 0, 255))


def generate_parse_table():
    try:
        parse_table = LL.ParseTable()
        path_parse = dpg.get_value('file_from_grammar')
        parse_table.generate_parse_table(path_parse)
        dpg.configure_item('result_generating',
                           default_value=f"Parse table was successfully generated to file: {path_parse}.xlsx",
                           color=(0, 255, 0, 255))
    except BaseException as err:
        dpg.configure_item('result_generating', default_value=f"Exception occurred during table generation:\n{err}",
                           color=(255, 0, 0, 255))


def initialize_lr3():
    with dpg.window(label="Лабораторная работа #3", tag='lr3', show=True, autosize=True, min_size=(1000, 800),
                    pos=(480, 0), on_close=lambda: dpg.delete_item('lr3')):
        initialize()
        with dpg.child_window(before='select_file_grammar', autosize_x=True, height=250):
            dpg.add_text('Generate new table with grammar')
            with dpg.group(horizontal=True):
                dpg.add_input_text(tag='file_from_grammar', default_value='lr3/tests/LL')
                dpg.add_button(label='Select file with grammar', callback=select_path, user_data='file_from_grammar')
            dpg.add_text(tag='result_generating')
            dpg.add_button(label='Generate', callback=generate_parse_table)
        dpg.configure_item('file_grammar', default_value='lr3/tests/LL.xlsx')
        dpg.configure_item('input_file', default_value='lr3/tests/test.txt')
        dpg.set_value('Manually_text', value=" float   b2( int sint , float int_, char ds[2], long double int7)  ; ")
        dpg.add_button(label="Analyze", callback=main, show=False, tag='continue')


def get_input_data():
    if dpg.get_value('input_method') == 'File':
        file_path = dpg.get_value('input_file')
        file = open(file_path, 'r')
        input_data = ''.join(file.readlines())
        file.close()
        return input_data
    else:
        return dpg.get_value('Manually_text')
