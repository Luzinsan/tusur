import openpyxl

is_nonterm = lambda s: s[0].isupper()

file_path_LL: str = "tests/test3"
example_string = "begin d comma d semi s comma s end ⊥"

dict_LL = dict()
wb = openpyxl.Workbook()
list_ods: openpyxl.Workbook = wb.create_sheet("List1", 0)
list_ods.append(('ПРАВИЛО', "START0"))

with open(file_path_LL) as file:
    index = 0
    for row in file.readlines():
        if len(list_splitted := row.split(" ->")) == 2:
            left, right = list_splitted
            rules = {}
            for rule in right.split('|'):
                rule = rule.strip()
                rules.update({index: rule})
                state0 = "∅" if is_nonterm(rule) else rule.split(" ")[0]
                list_ods.append((f"{left} -> {rule}", state0))
                index += 1
            dict_LL[left] = rules
            # print(rules)


# print(dict_LL)


def is_equal_cols(rows: int, worksheet: openpyxl.Workbook, compare_col1: int, compare_col2: int) -> bool:
    for row in range(2, rows + 1):
        # print(f"compare: {worksheet.cell(row=row, column=compare_col1).value} with { worksheet.cell(row=row, column=compare_col2).value}")
        if worksheet.cell(row=row, column=compare_col1).value != worksheet.cell(row=row, column=compare_col2).value:
            return False
    # print("Columns is equal!")
    return True


start_col = 2
num_rows_sheet = 0
while True:
    list_ods.cell(row=1, column=start_col + 1, value=f"START{start_col - 1}")
    # print(f"-----------------------ITERATION{col-2}:------------------------")
    for key in dict_LL.keys():
        # print(f"\n\t\tkey:{key}\tvalues:{dict_LL[key]}")
        for index, rule in dict_LL[key].items():
            new_starts = []
            if is_nonterm(rule):
                key_start = rule.split(" ")[0]
                # print(f"key_start:{key_start}")
                # print(f"from rules: {dict_LL[key_start]}")
                for start_index_row in dict_LL[key_start].keys():
                    # print("start_index: ", start_index_row)
                    new_starts.append(list_ods.cell(row=start_index_row + 2, column=start_col).value)
                    # print("new_starts: ", new_starts)
                new_starts = list(filter('∅'.__ne__, new_starts))
                if len(new_starts) == 0:
                    new_starts.append('∅')
            else:
                new_starts = [list_ods.cell(row=index + 2, column=start_col).value]

            list_ods.cell(row=index + 2, column=start_col + 1, value=" ".join(new_starts))
            num_rows_sheet = index + 2
            # print("\t\t\tResulting new_starts: ", new_starts)
    if is_equal_cols(num_rows_sheet, list_ods, start_col, start_col + 1):
        break
    start_col += 1


def review_next(key, nodes, index_node, follows, base_col):
    try:
        next_node = nodes[index_node + 1]
        # print("next node: ", next_node)
        # итерация по правилам следующего узла для поиска его символов-предсшественников
        for start_index_row in dict_LL[next_node].keys():
            # print("start_index: ", start_index_row)
            start_terms: list = list_ods.cell(row=start_index_row + 2, column=start_col).value.split(" ")
            # print("start_terms: ", start_terms)
            if 'e' in start_terms:
                start_terms.remove('e')
                follows += review_next(key, nodes, index_node + 1, follows, base_col)
                # follows += list_ods.cell(row=min(dict_LL[next_node].keys()) + 2, column=base_col).value.split(" ")
                # print("1: new_starts after replacing 'e': ", follows)
            else:
                follows += start_terms
                # print("2: new_starts: ", start_terms)
    except IndexError as _:
        # print("index_error: ", _)
        # print(f"append: row={min(dict_LL[key].keys()) + 2}\tcolumn={base_col}")
        follow_terms = list_ods.cell(row=min(dict_LL[key].keys()) + 2, column=base_col).value.split(" ")
        follows += follow_terms
        # print("new_follows: ", follows)
    except KeyError as _:
        # print("key_error:", _)
        follows.append(next_node)
    return list(set(follows))


def put_follows(col: int):
    list_ods.cell(row=1, column=col, value=f"FOLLOWS{col - start_col - 1}")
    # итерация по нетерминалам
    for key in dict_LL.keys():
        # print(f"\n\t\tkey:{key}\tvalues:{dict_LL[key]}")
        # итерация по правилам нетерминала
        for index_rule, rule in dict_LL[key].items():
            nodes = rule.split(" ")
            # итерация по терминальным и нетерминальным узлам
            for index_node, node in enumerate(nodes):
                # для каждого нетерминала определяем последующие узлы
                if is_nonterm(node):
                    # print(node, " - ", dict_LL[node])
                    old_follows = list_ods.cell(row=min(dict_LL[node].keys()) + 2, column=col).value
                    follows = old_follows.split(" ") if old_follows else []
                    # print("Already having follows: ", follows)
                    follows = review_next(key, nodes, index_node, follows, col)
                    # print("Finale FOLLOWS: ", follows)
                    # print("PUT FOLLOWS AT: row=", min(dict_LL[node].keys()) + 2, " column=", col)
                    list_ods.cell(row=min(dict_LL[node].keys()) + 2, column=col, value=" ".join(follows))
    return list_ods


start_col += 1
follow_col = start_col + 1
for key in dict_LL.keys():
    list_ods.merge_cells(start_row=min(dict_LL[key].keys()) + 2, start_column=follow_col,
                         end_row=max(dict_LL[key].keys()) + 2, end_column=follow_col)
list_ods.cell(row=2, column=follow_col, value="⊥")

list_ods = put_follows(follow_col)
follow_col += 1
while True:
    for key in dict_LL.keys():
        list_ods.merge_cells(start_row=min(dict_LL[key].keys()) + 2, start_column=follow_col,
                             end_row=max(dict_LL[key].keys()) + 2, end_column=follow_col)
        old_value = list_ods.cell(row=min(dict_LL[key].keys()) + 2, column=follow_col - 1).value
        list_ods.cell(row=min(dict_LL[key].keys()) + 2, column=follow_col, value=old_value)
    list_ods = put_follows(follow_col)
    if is_equal_cols(num_rows_sheet, list_ods, follow_col - 1, follow_col):
        break
    follow_col += 1


def next_starts(key, starts, nodes, index_node):
    try:
        next_node = nodes[index_node + 1]
        # print("next node: ", next_node)
        for start_index_row in dict_LL[next_node].keys():
            # print("start_index: ", start_index_row)
            starts += list_ods.cell(row=start_index_row + 2, column=start_col).value.split(" ")
            if 'e' in starts:
                starts.remove("e")
                starts = next_starts(key, starts, nodes, index_node + 1)
            # print("new_starts: ", starts)
    except IndexError as _:
        # print("index_error: ", _)
        # print(f"append: row={min(dict_LL[key].keys()) + 2}\tcolumn={follow_col}")
        follow_terms = list_ods.cell(row=min(dict_LL[key].keys()) + 2, column=follow_col).value.split(" ")
        starts += follow_terms
        # print("new_starts: ", follow_terms)
    except KeyError as _:
        # print("key_error:", _)
        starts.append(next_node)
    return starts


def is_last_node(current_index, rule_nodes):
    try:
        if rule_nodes[current_index + 1]:
            return False
    except IndexError as _:
        return True


list_ods.cell(row=1, column=follow_col + 1, value="DIRECTIONS")
parse_table: openpyxl.Workbook = wb.create_sheet("parse_table", 1)
parse_table.append(('НЕТЕРМИНАЛЫ', "terminals", "jump", "accept", "stack", "return", "error"))
index_term = 1
dict_M = dict()
for key in dict_LL.keys():
    # print(f"\n\t\tkey:{key}\tvalues:{dict_LL[key]}")
    for index in dict_LL[key].keys():
        index_term += 1
        # print("left term: ", key, "\tindex left term: ", index_term - 1)
        direction_terms: list = list_ods.cell(row=index + 2, column=start_col).value.split(" ")
        # print("start_terms: ", direction_terms)
        if 'e' in direction_terms:
            direction_terms.remove("e")
            follow_e = list_ods.cell(row=min(dict_LL[key].keys()) + 2, column=follow_col).value.split(" ")
            # print("replacing E: ", follow_e)
            direction_terms += follow_e
        list_ods.cell(row=index + 2, column=follow_col + 1, value=" ".join(direction_terms))
        dict_M.update({(key, index_term): {}})
        # print("dict_M: ", dict_M)
        parse_table.append(("left: " + key, " ".join(direction_terms), "", "False", "False", "False", "False"))
    parse_table.cell(row=index_term, column=7, value="True")
    last_left_term = index_term - 1
    # поиск направляющих символов в правой части правил
    # B -> aACg
    # T(Ai) = S(A)
    # if 'e' in S(A)
    #     remove 'e'
    #     + S(C)

    for index_rule, rule in dict_LL[key].items():
        # print("Current key: ", key, "\tindex rule: ", index_rule)
        nodes = rule.split(" ")
        # итерация по терминальным и нетерминальным узлам
        for index_node, node in enumerate(nodes):
            index_term += 1
            # print("right term: ", node, "\tindex right term: ", index_term - 1)
            # для каждого нетерминала определяем направляющие узлы
            accept = "False"
            stack = "False"
            if is_nonterm(node):
                stack = str(not is_last_node(index_node, nodes))
                starts = []
                for start_index_row in dict_LL[node].keys():
                    # print("start_index: ", start_index_row)
                    starts += list_ods.cell(row=start_index_row + 2, column=start_col).value.split(" ")
                    # print("starts: ", starts)
                    if 'e' in starts:
                        starts.remove("e")
                        starts = next_starts(key, starts, nodes, index_node)
            elif node == 'e':
                starts = list_ods.cell(row=min(dict_LL[key].keys()) + 2, column=follow_col).value.split(" ")
            else:
                starts = [node]
                accept = "True"
            global_index_rule = index_rule - min(dict_LL[key].keys())
            # print("number of rules: ", len(dict_LL[key]), "\tglobal index rule: ", global_index_rule, "\tindex of term: ", index_term - 1)
            dict_M[(key, last_left_term - len(dict_LL[key]) + 2 + global_index_rule)].update({index_term: node})
            # print("dict_M: ", dict_M)
            # print("FINALE starts: ", starts)
            parse_table.append(("right: " + node, " ".join(starts), "", accept, stack, "False", "True"))

# jumps
for key, values in dict_M.items():
    # print("key: ", key, "\tvalue", values)
    parse_table.cell(row=key[1], column=3, value=min(values.keys()))
    for index_node, node in values.items():
        if is_nonterm(node):
            # print(f"loop\t\tindex_node={index_node}\tnode={node}")
            for root_key in dict_M.keys():
                # print(f"loop\t\troot_key={root_key}\tnode={node}")
                if root_key[0] == node:
                    # print(f"root_key[0]={root_key[0]}")
                    parse_table.cell(row=index_node, column=3, value=root_key[1])
                    break
        else:
            try:
                # print(f"try: index_node={index_node + 1}")
                next_node = values[index_node + 1]
                # print("success: ", next_node)
                parse_table.cell(row=index_node, column=3, value=index_node + 1)
            except KeyError as _:
                # print("is last node -> 0")
                parse_table.cell(row=index_node, column=3, value=0)
                parse_table.cell(row=index_node, column=6, value="True")

wb.save('LL.xlsx')


M_stack = [0]
i = 2
k = 0
example_string = example_string.split(" ")
while True:
    print("\ncurrent i: ", i-1)
    print("current token: ", example_string[k], "\tk=", k)
    print("terminals: ", parse_table.cell(row=i, column=2).value.split(" "))
    if example_string[k] in parse_table.cell(row=i, column=2).value.split(" "):
        print("accepts: ", parse_table.cell(row=i, column=4).value)
        print("stack: ", parse_table.cell(row=i, column=5).value)
        print("returns: ", parse_table.cell(row=i, column=6).value)
        print("jumps: ", parse_table.cell(row=i, column=3).value - 1)
        if parse_table.cell(row=i, column=4).value == "True":
            k += 1
        if parse_table.cell(row=i, column=5).value == "True":
            M_stack.append(i + 1)
        if parse_table.cell(row=i, column=6).value == "True":
            i = M_stack.pop()
            if i == 0:
                break
            else:
                continue
        else:  # parse_table.cell(row=i, column=3).value != 0:
            i = parse_table.cell(row=i, column=3).value
        print("Stack: ", M_stack)
    elif parse_table.cell(row=i, column=7).value == "False":
        i += 1
    else:
        print("error: ", parse_table.cell(row=i, column=7).value)
        break

print("\t\tlength of stack: ", len(M_stack), "\tstack:", M_stack, "\t string[k]:", example_string[k])
if len(M_stack) == 0 and example_string[k] == '⊥':
    print("SUCCESS PARSED!")
else:
    print("FAILED PARSED! at: ", k, " - ", example_string[k])

